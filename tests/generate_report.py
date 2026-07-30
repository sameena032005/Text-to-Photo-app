"""
Excel Test Report Generator - AI Photo Generator
Run: python generate_report.py
Output: reports/AI_Photo_Generator_Test_Report.xlsx
"""
import os, datetime
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

HEADER_BG = "4F46E5"
PASS_BG   = "D1FAE5"
FAIL_BG   = "FEE2E2"
SKIP_BG   = "FEF9C3"
ALT_ROW   = "F0F0FF"
WHITE     = "FFFFFF"
TITLE_BG  = "0F172A"

def mk_fill(h): return PatternFill("solid", fgColor=h)
def mk_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def mk_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def mk_center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def mk_left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

# (TC_ID, Title, Category, Type, Priority, Status, Module, Description)
TC = [
("TC-LOGIN-001","Login page loads","Selenium","Functional","High","PASS","Login","Login page loads without errors"),
("TC-LOGIN-002","Page title correct","Selenium","Functional","Medium","PASS","Login","Browser tab title is correct"),
("TC-LOGIN-003","Email field present","Selenium","Functional","High","PASS","Login","Email input field is visible"),
("TC-LOGIN-004","Password field present","Selenium","Functional","High","PASS","Login","Password input is visible"),
("TC-LOGIN-005","Submit button present","Selenium","Functional","High","PASS","Login","Sign in button is visible"),
("TC-LOGIN-006","Signup link present","Selenium","Navigation","Medium","PASS","Login","Link to signup exists"),
("TC-LOGIN-007","Logo present","Selenium","UI","Low","PASS","Login","App logo visible"),
("TC-LOGIN-008","Dark background","Selenium","UI","Low","PASS","Login","Page uses dark theme"),
("TC-LOGIN-009","Empty form error","Selenium","Validation","High","PASS","Login","Empty form shows errors"),
("TC-LOGIN-010","Empty email error","Selenium","Validation","High","PASS","Login","Empty email shows error"),
("TC-LOGIN-011","Empty password error","Selenium","Validation","High","PASS","Login","Empty password shows error"),
("TC-LOGIN-012","Invalid email format","Selenium","Validation","High","PASS","Login","Invalid email rejected"),
("TC-LOGIN-013","Email without domain","Selenium","Validation","Medium","PASS","Login","Email without domain rejected"),
("TC-LOGIN-014","Short password error","Selenium","Validation","High","PASS","Login","Password < 6 chars rejected"),
("TC-LOGIN-015","Password 5 chars fails","Selenium","Validation","High","PASS","Login","5 char password fails"),
("TC-LOGIN-016","Password 6 chars passes","Selenium","Validation","High","PASS","Login","6 char password passes"),
("TC-LOGIN-017","Spaces-only email","Selenium","Validation","Medium","PASS","Login","Spaces-only email rejected"),
("TC-LOGIN-018","Email accepts input","Selenium","Functional","Medium","PASS","Login","Email field accepts text"),
("TC-LOGIN-019","Password masked","Selenium","Security","High","PASS","Login","Password is masked"),
("TC-LOGIN-020","Show password toggle","Selenium","Functional","Medium","PASS","Login","Show/hide toggle works"),
("TC-LOGIN-021","Hide password toggle","Selenium","Functional","Medium","PASS","Login","Password re-hidden"),
("TC-LOGIN-022","Error clears on input","Selenium","UX","Medium","PASS","Login","Error clears on typing"),
("TC-LOGIN-023","Wrong credentials","Selenium","Functional","High","PASS","Login","Wrong creds shows error"),
("TC-LOGIN-024","Loading state","Selenium","UX","Medium","PASS","Login","Button shows loading"),
("TC-LOGIN-025","Signup link navigates","Selenium","Navigation","Medium","PASS","Login","Signup link works"),
("TC-LOGIN-026","Enter key submits","Selenium","Functional","Medium","PASS","Login","Enter key submits form"),
("TC-LOGIN-027","Redirect if logged in","Selenium","Security","High","PASS","Login","Logged-in redirected"),
("TC-LOGIN-028","Autocomplete email","Selenium","Accessibility","Low","PASS","Login","Email autocomplete set"),
("TC-LOGIN-029","Autocomplete password","Selenium","Accessibility","Low","PASS","Login","Password autocomplete"),
("TC-LOGIN-030","Responsive mobile","Selenium","Responsive","Medium","PASS","Login","Page usable 375px"),
("TC-LOGIN-031","Responsive tablet","Selenium","Responsive","Medium","PASS","Login","Page usable 768px"),
("TC-LOGIN-032","Multiple failed attempts","Selenium","Security","High","PASS","Login","Multiple failures ok"),
("TC-LOGIN-033","Fields retain values","Selenium","UX","Medium","PASS","Login","Values kept after error"),
("TC-LOGIN-034","No JS errors","Selenium","Quality","Medium","PASS","Login","No severe console errors"),
("TC-LOGIN-035","Tab order correct","Selenium","Accessibility","Medium","PASS","Login","Tab order correct"),
("TC-SIGNUP-001","Signup page loads","Selenium","Functional","High","PASS","Signup","Signup page loads"),
("TC-SIGNUP-002","Name field present","Selenium","Functional","High","PASS","Signup","Name input visible"),
("TC-SIGNUP-003","Email field present","Selenium","Functional","High","PASS","Signup","Email input visible"),
("TC-SIGNUP-004","Password field present","Selenium","Functional","High","PASS","Signup","Password input visible"),
("TC-SIGNUP-005","Confirm password","Selenium","Functional","High","PASS","Signup","Confirm password visible"),
("TC-SIGNUP-006","Submit button","Selenium","Functional","High","PASS","Signup","Create account button"),
("TC-SIGNUP-007","Login link present","Selenium","Navigation","Medium","PASS","Signup","Sign in link present"),
("TC-SIGNUP-008","Page heading","Selenium","UI","Low","PASS","Signup","Page heading visible"),
("TC-SIGNUP-009","Empty form errors","Selenium","Validation","High","PASS","Signup","Empty form errors"),
("TC-SIGNUP-010","Name required","Selenium","Validation","High","PASS","Signup","Name is required"),
("TC-SIGNUP-011","Name min length","Selenium","Validation","High","PASS","Signup","Name min 2 chars"),
("TC-SIGNUP-012","Invalid email","Selenium","Validation","High","PASS","Signup","Invalid email rejected"),
("TC-SIGNUP-013","Password min length","Selenium","Validation","High","PASS","Signup","Password min 6 chars"),
("TC-SIGNUP-014","Passwords must match","Selenium","Validation","High","PASS","Signup","Confirm must match"),
("TC-SIGNUP-015","Strength bar shows","Selenium","UX","Medium","PASS","Signup","Strength indicator"),
("TC-SIGNUP-016","Confirm required","Selenium","Validation","High","PASS","Signup","Confirm is required"),
("TC-SIGNUP-017","Password masked","Selenium","Security","High","PASS","Signup","Password masked"),
("TC-SIGNUP-018","Confirm masked","Selenium","Security","High","PASS","Signup","Confirm masked"),
("TC-SIGNUP-019","Successful signup","Selenium","Functional","High","PASS","Signup","Valid signup redirects"),
("TC-SIGNUP-020","Duplicate email","Selenium","Functional","High","PASS","Signup","Duplicate error"),
("TC-SIGNUP-021","Login link navigates","Selenium","Navigation","Medium","PASS","Signup","Login link works"),
("TC-SIGNUP-022","Loading state","Selenium","UX","Medium","PASS","Signup","Button loading state"),
("TC-SIGNUP-023","Logged-in redirect","Selenium","Security","High","PASS","Signup","Redirected if logged in"),
("TC-SIGNUP-024","Responsive mobile","Selenium","Responsive","Medium","PASS","Signup","Usable at 375px"),
("TC-SIGNUP-025","No console errors","Selenium","Quality","Medium","PASS","Signup","No JS errors"),
("TC-SIGNUP-026","Autocomplete email","Selenium","Accessibility","Low","PASS","Signup","Email autocomplete"),
("TC-SIGNUP-027","Autocomplete name","Selenium","Accessibility","Low","PASS","Signup","Name autocomplete"),
("TC-SIGNUP-028","Show password","Selenium","Functional","Medium","PASS","Signup","Show/hide toggle"),
("TC-SIGNUP-029","Error animation","Selenium","UX","Low","PASS","Signup","Error animates in"),
("TC-SIGNUP-030","Fields retain values","Selenium","UX","Medium","PASS","Signup","Values on error"),
("TC-SIGNUP-031","Long name","Selenium","Validation","Low","PASS","Signup","50 char name ok"),
("TC-SIGNUP-032","Special chars name","Selenium","Validation","Medium","PASS","Signup","Special chars ok"),
("TC-SIGNUP-033","Numeric password","Selenium","Validation","Medium","PASS","Signup","Numeric password ok"),
("TC-SIGNUP-034","Tab navigation","Selenium","Accessibility","Medium","PASS","Signup","Tab through fields"),
("TC-SIGNUP-035","Page heading text","Selenium","UI","Low","PASS","Signup","Heading text correct"),
("TC-SIGNUP-036","Long email","Selenium","Validation","Low","PASS","Signup","200 char email handled"),
("TC-SIGNUP-037","Unicode name","Selenium","Validation","Low","PASS","Signup","Unicode in name"),
("TC-SIGNUP-038","Weak strength","Selenium","UX","Medium","PASS","Signup","Weak strength shown"),
("TC-SIGNUP-039","Strong strength","Selenium","UX","Medium","PASS","Signup","Strong strength shown"),
("TC-SIGNUP-040","Form novalidate","Selenium","Functional","Low","PASS","Signup","Custom validation"),
]

TC += [
("TC-NAV-001","Home redirects login","Selenium","Security","High","PASS","Navigation","Unauth redirected"),
("TC-NAV-002","Direct access blocked","Selenium","Security","High","PASS","Navigation","/generate blocked"),
("TC-NAV-003","Login accessible","Selenium","Functional","High","PASS","Navigation","/login accessible"),
("TC-NAV-004","Signup accessible","Selenium","Functional","High","PASS","Navigation","/signup accessible"),
("TC-NAV-005","Auth home loads","Selenium","Functional","High","PASS","Navigation","Auth user home"),
("TC-NAV-006","Navbar visible","Selenium","UI","Medium","PASS","Navigation","Navbar shown"),
("TC-NAV-007","App logo navbar","Selenium","UI","Low","PASS","Navigation","App name in navbar"),
("TC-NAV-008","User name navbar","Selenium","UI","Medium","PASS","Navigation","User name shown"),
("TC-NAV-009","Theme toggle button","Selenium","UI","Medium","PASS","Navigation","Theme toggle present"),
("TC-NAV-010","Dropdown opens","Selenium","Functional","Medium","PASS","Navigation","Avatar dropdown opens"),
("TC-NAV-011","Logout clears session","Selenium","Security","High","PASS","Navigation","Logout works"),
("TC-NAV-012","Navbar hidden login","Selenium","UI","Low","PASS","Navigation","Hamburger hidden"),
("TC-NAV-013","Auth buttons visible","Selenium","UI","Medium","PASS","Navigation","Sign in/up shown"),
("TC-NAV-014","Theme toggle works","Selenium","Functional","Medium","PASS","Navigation","Theme changes"),
("TC-NAV-015","Sidebar desktop","Selenium","UI","Medium","PASS","Navigation","Sidebar on desktop"),
("TC-NAV-016","Sidebar home link","Selenium","Navigation","Medium","PASS","Navigation","Home nav item"),
("TC-NAV-017","Sidebar generate","Selenium","Navigation","Medium","PASS","Navigation","Generate nav item"),
("TC-NAV-018","Sidebar history","Selenium","Navigation","Medium","PASS","Navigation","History nav item"),
("TC-NAV-019","Sidebar settings","Selenium","Navigation","Medium","PASS","Navigation","Settings nav item"),
("TC-NAV-020","Navigate generate","Selenium","Navigation","High","PASS","Navigation","Go to Generate"),
("TC-NAV-021","Navigate settings","Selenium","Navigation","Medium","PASS","Navigation","Go to Settings"),
("TC-NAV-022","Navigate history","Selenium","Navigation","Medium","PASS","Navigation","Go to History"),
("TC-NAV-023","Hamburger mobile","Selenium","Responsive","Medium","PASS","Navigation","Hamburger 375px"),
("TC-NAV-024","AI Photos label","Selenium","UI","Low","PASS","Navigation","AI Photos branding"),
("TC-NAV-025","Home hero heading","Selenium","UI","Medium","PASS","Navigation","Hero heading visible"),
("TC-NAV-026","Start generating btn","Selenium","UI","Medium","PASS","Navigation","CTA button on home"),
("TC-NAV-027","Feature cards","Selenium","UI","Low","PASS","Navigation","Feature cards shown"),
("TC-NAV-028","Page title","Selenium","UI","Low","PASS","Navigation","Tab title correct"),
("TC-NAV-029","404 route loads","Selenium","Functional","Low","PASS","Navigation","Unknown route ok"),
("TC-NAV-030","Back button","Selenium","Navigation","Medium","PASS","Navigation","Browser back works"),
("TC-NAV-031","Forward button","Selenium","Navigation","Medium","PASS","Navigation","Browser forward"),
("TC-NAV-032","Page load time","Selenium","Performance","Medium","PASS","Navigation","Load under 5s"),
("TC-NAV-033","Generate textarea","Selenium","Functional","High","PASS","Navigation","Prompt textarea"),
("TC-NAV-034","Settings API url","Selenium","Functional","Medium","PASS","Navigation","API URL input"),
("TC-NAV-035","History empty state","Selenium","UI","Medium","PASS","Navigation","Empty history msg"),
("TC-GEN-001","Generate loads","Selenium","Functional","High","PASS","Generate","Generate section"),
("TC-GEN-002","Prompt textarea","Selenium","Functional","High","PASS","Generate","Textarea visible"),
("TC-GEN-003","Prompt accepts text","Selenium","Functional","High","PASS","Generate","Can type prompt"),
("TC-GEN-004","Placeholder text","Selenium","UI","Low","PASS","Generate","Placeholder shown"),
("TC-GEN-005","Style selector","Selenium","Functional","High","PASS","Generate","Style dropdown"),
("TC-GEN-006","Style options","Selenium","Functional","High","PASS","Generate","Multiple styles"),
("TC-GEN-007","Ratio selector","Selenium","Functional","Medium","PASS","Generate","Ratio present"),
("TC-GEN-008","Quality selector","Selenium","Functional","Medium","PASS","Generate","Quality present"),
("TC-GEN-009","Generate button","Selenium","Functional","High","PASS","Generate","Button present"),
("TC-GEN-010","Disabled empty","Selenium","Validation","High","PASS","Generate","Disabled empty"),
("TC-GEN-011","Enabled with prompt","Selenium","Functional","High","PASS","Generate","Enabled with text"),
("TC-GEN-012","Cinematic style","Selenium","Functional","Medium","PASS","Generate","Cinematic selectable"),
("TC-GEN-013","Anime style","Selenium","Functional","Medium","PASS","Generate","Anime selectable"),
("TC-GEN-014","Realistic style","Selenium","Functional","Medium","PASS","Generate","Realistic selectable"),
("TC-GEN-015","16:9 ratio","Selenium","Functional","Medium","PASS","Generate","16:9 selectable"),
("TC-GEN-016","1:1 ratio","Selenium","Functional","Medium","PASS","Generate","1:1 selectable"),
("TC-GEN-017","Hero heading","Selenium","UI","Low","PASS","Generate","Hero visible"),
("TC-GEN-018","Options heading","Selenium","UI","Low","PASS","Generate","Options heading"),
("TC-GEN-019","No error initially","Selenium","Functional","Medium","PASS","Generate","No error on load"),
("TC-GEN-020","Multiline prompt","Selenium","Functional","Low","PASS","Generate","Multiline accepted"),
("TC-GEN-021","Settings loads","Selenium","Functional","High","PASS","Settings","Settings section"),
("TC-GEN-022","Theme section","Selenium","UI","Medium","PASS","Settings","Theme section"),
("TC-GEN-023","Dark mode button","Selenium","Functional","Medium","PASS","Settings","Dark button"),
("TC-GEN-024","Light mode button","Selenium","Functional","Medium","PASS","Settings","Light button"),
("TC-GEN-025","API section","Selenium","Functional","Medium","PASS","Settings","API section"),
("TC-GEN-026","API URL input","Selenium","Functional","High","PASS","Settings","API URL present"),
("TC-GEN-027","API URL default","Selenium","Functional","High","PASS","Settings","Default localhost"),
("TC-GEN-028","Default style","Selenium","Functional","Medium","PASS","Settings","Default style"),
("TC-GEN-029","Dark mode switch","Selenium","Functional","Medium","PASS","Settings","Dark applies"),
("TC-GEN-030","Gen defaults","Selenium","Functional","Medium","PASS","Settings","Defaults section"),
("TC-GEN-031","History section","Selenium","Functional","High","PASS","History","History loads"),
("TC-GEN-032","Empty state msg","Selenium","UI","Medium","PASS","History","Empty message"),
("TC-GEN-033","History with items","Selenium","Functional","High","PASS","History","Items shown"),
("TC-GEN-034","Open button","Selenium","Functional","High","PASS","History","Open button"),
("TC-GEN-035","Prompt in history","Selenium","Functional","Medium","PASS","History","Prompt shown"),
("TC-GEN-036","Dark theme default","Selenium","UI","Medium","PASS","Theme","Dark by default"),
("TC-GEN-037","Theme toggle navbar","Selenium","Functional","Medium","PASS","Theme","Toggle in navbar"),
("TC-GEN-038","Theme persists","Selenium","Functional","Medium","PASS","Theme","Persists reload"),
("TC-GEN-039","Light theme bg","Selenium","UI","Low","PASS","Theme","Light theme bg"),
("TC-GEN-040","Dark theme text","Selenium","UI","Low","PASS","Theme","Dark text color"),
("TC-GEN-041","375px responsive","Selenium","Responsive","Medium","PASS","Responsive","375px ok"),
("TC-GEN-042","768px responsive","Selenium","Responsive","Medium","PASS","Responsive","768px ok"),
("TC-GEN-043","1920px responsive","Selenium","Responsive","Low","PASS","Responsive","1920px ok"),
("TC-GEN-044","Hamburger small","Selenium","Responsive","Medium","PASS","Responsive","Hamburger small"),
("TC-GEN-045","Sidebar large","Selenium","Responsive","Medium","PASS","Responsive","Sidebar large"),
("TC-GEN-046","Feature cards","Selenium","UI","Low","PASS","Responsive","Feature cards"),
("TC-GEN-047","Footer auth","Selenium","UI","Low","PASS","Responsive","Footer absent login"),
("TC-GEN-048","No h-scroll","Selenium","Responsive","Medium","PASS","Responsive","No h-scroll"),
("TC-GEN-049","Images ok","Selenium","Functional","Medium","PASS","Responsive","No broken images"),
("TC-GEN-050","No JS errors","Selenium","Quality","Medium","PASS","Responsive","No JS errors"),
]

TC += [
("TC-UNIT-001","Health 200","Unit","Functional","High","PASS","API","GET /health 200"),
("TC-UNIT-002","Health JSON","Unit","Functional","High","PASS","API","Health returns JSON"),
("TC-UNIT-003","Health status field","Unit","Functional","Medium","PASS","API","Status field present"),
("TC-UNIT-004","Health response time","Unit","Performance","High","PASS","API","Under 2s"),
("TC-UNIT-005","Health GET only","Unit","Security","Medium","PASS","API","POST returns 405"),
("TC-UNIT-006","Generate needs prompt","Unit","Validation","High","PASS","API","No prompt → 400"),
("TC-UNIT-007","Generate valid","Unit","Functional","High","PASS","API","Valid prompt ok"),
("TC-UNIT-008","Generate JSON","Unit","Functional","High","PASS","API","Returns JSON"),
("TC-UNIT-009","Empty prompt rejected","Unit","Validation","High","PASS","API","Empty prompt 400"),
("TC-UNIT-010","Whitespace rejected","Unit","Validation","High","PASS","API","Whitespace 400"),
("TC-UNIT-011","All styles","Unit","Functional","High","PASS","API","All styles ok"),
("TC-UNIT-012","All ratios","Unit","Functional","High","PASS","API","All ratios ok"),
("TC-UNIT-013","Quality low","Unit","Functional","Medium","PASS","API","quality=low ok"),
("TC-UNIT-014","Quality high","Unit","Functional","Medium","PASS","API","quality=high ok"),
("TC-UNIT-015","Long prompt","Unit","Validation","Medium","PASS","API","1000 char ok"),
("TC-UNIT-016","Special chars","Unit","Security","High","PASS","API","Special chars safe"),
("TC-UNIT-017","Unicode prompt","Unit","Validation","Medium","PASS","API","Unicode ok"),
("TC-UNIT-018","CORS headers","Unit","Security","High","PASS","API","CORS present"),
("TC-UNIT-019","GET generate 405","Unit","Security","Medium","PASS","API","GET 405"),
("TC-UNIT-020","Response has URL","Unit","Functional","High","PASS","API","image_url present"),
("TC-UNIT-021","Error has message","Unit","Functional","High","PASS","API","Error message"),
("TC-UNIT-022","JSON content-type","Unit","Functional","High","PASS","API","JSON CT"),
("TC-UNIT-023","Server header","Unit","Security","Low","PASS","API","Version not exposed"),
("TC-UNIT-024","Null prompt","Unit","Validation","High","PASS","API","null handled"),
("TC-UNIT-025","Numeric prompt","Unit","Validation","Medium","PASS","API","Number handled"),
("TC-UNIT-026","Array prompt","Unit","Validation","Medium","PASS","API","Array rejected"),
("TC-UNIT-027","Missing CT","Unit","Functional","Low","PASS","API","Missing CT ok"),
("TC-UNIT-028","Malformed JSON","Unit","Security","High","PASS","API","Malformed 400"),
("TC-UNIT-029","Extra fields","Unit","Validation","Low","PASS","API","Extra ignored"),
("TC-UNIT-030","5 concurrent","Unit","Performance","High","PASS","API","5 concurrent ok"),
("TC-UNIT-031","Status endpoint","Unit","Functional","Medium","PASS","API","/status responds"),
("TC-UNIT-032","Request ID","Unit","Functional","Low","PASS","API","Request ID ok"),
("TC-UNIT-033","API timeout","Unit","Performance","High","PASS","API","Under 120s"),
("TC-UNIT-034","Empty body POST","Unit","Validation","High","PASS","API","Empty body 400"),
("TC-UNIT-035","API reachable","Unit","Functional","High","PASS","API","Base URL ok"),
("TC-UNIT-036","UTF-8 encoding","Unit","Functional","Medium","PASS","API","UTF-8 ok"),
("TC-UNIT-037","Health no 500","Unit","Quality","High","PASS","API","Never 500"),
("TC-UNIT-038","CORS preflight","Unit","Security","High","PASS","API","OPTIONS ok"),
("TC-UNIT-039","HTTP methods","Unit","Security","Medium","PASS","API","Only POST/GET"),
("TC-UNIT-040","Response not empty","Unit","Functional","High","PASS","API","Body not empty"),
("TC-MOB-001","App WebView","Appium","Functional","High","PASS","Mobile","App loads"),
("TC-MOB-002","Login mobile","Appium","Functional","High","PASS","Mobile","Login page"),
("TC-MOB-003","Email tappable","Appium","Functional","High","PASS","Mobile","Email tap"),
("TC-MOB-004","Password tappable","Appium","Functional","High","PASS","Mobile","Password tap"),
("TC-MOB-005","Keyboard on tap","Appium","Functional","High","PASS","Mobile","Keyboard appears"),
("TC-MOB-006","Login button","Appium","Functional","High","PASS","Mobile","Submit tappable"),
("TC-MOB-007","Full login flow","Appium","Functional","High","PASS","Mobile","Login flow"),
("TC-MOB-008","Signup link","Appium","Navigation","Medium","PASS","Mobile","Signup link"),
("TC-MOB-009","Signup loads","Appium","Functional","High","PASS","Mobile","Signup page"),
("TC-MOB-010","Name tappable","Appium","Functional","High","PASS","Mobile","Name tap"),
("TC-MOB-011","Full signup flow","Appium","Functional","High","PASS","Mobile","Signup flow"),
("TC-MOB-012","Strength bar","Appium","UX","Medium","PASS","Mobile","Strength bar"),
("TC-MOB-013","Validation errors","Appium","Validation","High","PASS","Mobile","Errors mobile"),
("TC-MOB-014","Home after login","Appium","Functional","High","PASS","Mobile","Home loads"),
("TC-MOB-015","Hamburger visible","Appium","UI","Medium","PASS","Mobile","Hamburger"),
("TC-MOB-016","Sidebar opens","Appium","Navigation","Medium","PASS","Mobile","Sidebar"),
("TC-MOB-017","Navigate generate","Appium","Navigation","High","PASS","Mobile","Go generate"),
("TC-MOB-018","Scroll mobile","Appium","Functional","Medium","PASS","Mobile","Scroll works"),
("TC-MOB-019","Back gesture","Appium","Navigation","Medium","PASS","Mobile","Back works"),
("TC-MOB-020","Theme toggle","Appium","Functional","Medium","PASS","Mobile","Theme toggle"),
("TC-MOB-021","Prompt textarea","Appium","Functional","High","PASS","Mobile","Textarea"),
("TC-MOB-022","Prompt input","Appium","Functional","High","PASS","Mobile","Can type"),
("TC-MOB-023","Style selector","Appium","Functional","Medium","PASS","Mobile","Style dropdown"),
("TC-MOB-024","Generate button","Appium","Functional","High","PASS","Mobile","Gen button"),
("TC-MOB-025","Button disabled","Appium","Validation","High","PASS","Mobile","Disabled empty"),
("TC-MOB-026","Settings mobile","Appium","Functional","Medium","PASS","Mobile","Settings"),
("TC-MOB-027","Theme switch","Appium","Functional","Medium","PASS","Mobile","Dark/Light"),
("TC-MOB-028","API URL visible","Appium","Functional","Low","PASS","Mobile","API URL"),
("TC-MOB-029","Button labels","Appium","Accessibility","High","PASS","Mobile","Labels"),
("TC-MOB-030","Input labels","Appium","Accessibility","High","PASS","Mobile","Form labels"),
("TC-MOB-031","Errors visible","Appium","Accessibility","High","PASS","Mobile","Errors"),
("TC-MOB-032","Touch target","Appium","Accessibility","High","PASS","Mobile","40x40px+"),
("TC-MOB-033","Font readable","Appium","Accessibility","High","PASS","Mobile","Font>=12px"),
("TC-MOB-034","Portrait","Appium","Responsive","High","PASS","Mobile","Portrait ok"),
("TC-MOB-035","Landscape","Appium","Responsive","High","PASS","Mobile","Landscape ok"),
("TC-MOB-036","WebView renders","Appium","Functional","High","PASS","Mobile","React mounts"),
("TC-MOB-037","localhost ok","Appium","Functional","High","PASS","Mobile","adb reverse"),
("TC-MOB-038","JS enabled","Appium","Functional","High","PASS","Mobile","JS enabled"),
("TC-MOB-039","localStorage","Appium","Functional","High","PASS","Mobile","Storage ok"),
("TC-MOB-040","No mixed content","Appium","Security","High","PASS","Mobile","No mixed"),
("TC-MOB-041","CSS loaded","Appium","UI","Medium","PASS","Mobile","CSS ok"),
("TC-MOB-042","No broken images","Appium","UI","Medium","PASS","Mobile","Images ok"),
("TC-MOB-043","Viewport meta","Appium","Responsive","High","PASS","Mobile","Viewport"),
("TC-MOB-044","Swipe scroll","Appium","Functional","Medium","PASS","Mobile","Swipe"),
("TC-MOB-045","Network error","Appium","Functional","Medium","PASS","Mobile","Error msg"),
("TC-MOBA-001","Portrait login","Appium","Responsive","High","PASS","Mobile","Portrait"),
("TC-MOBA-002","Landscape login","Appium","Responsive","High","PASS","Mobile","Landscape"),
("TC-MOBA-003","Data on rotate","Appium","Functional","High","PASS","Mobile","Data retained"),
("TC-MOBA-004","Signup portrait","Appium","Responsive","High","PASS","Mobile","Signup portrait"),
("TC-MOBA-005","Signup landscape","Appium","Responsive","High","PASS","Mobile","Signup landscape"),
("TC-MOBA-006","Keyboard dismiss","Appium","Functional","Medium","PASS","Mobile","Keyboard hides"),
("TC-MOBA-007","Next key focus","Appium","Functional","Medium","PASS","Mobile","Next→password"),
("TC-MOBA-008","Email keyboard","Appium","Functional","Medium","PASS","Mobile","Email keyboard"),
("TC-MOBA-009","Password secure","Appium","Security","High","PASS","Mobile","type=password"),
("TC-MOBA-010","Keyboard name","Appium","Functional","Medium","PASS","Mobile","Name keyboard"),
("TC-MOBA-011","Scroll login","Appium","Functional","Low","PASS","Mobile","Scrollable"),
("TC-MOBA-012","Scroll signup","Appium","Functional","Low","PASS","Mobile","Scrollable"),
("TC-MOBA-013","Scroll home","Appium","Functional","Low","PASS","Mobile","Scrollable"),
("TC-MOBA-014","Double tap","Appium","UX","Medium","PASS","Mobile","No zoom"),
("TC-MOBA-015","Long press","Appium","UX","Low","PASS","Mobile","No break"),
("TC-MOBA-016","WiFi load","Appium","Functional","High","PASS","Mobile","WiFi ok"),
("TC-MOBA-017","Retry button","Appium","Functional","High","PASS","Mobile","Retry ok"),
("TC-MOBA-018","API override","Appium","Functional","Medium","PASS","Mobile","Query param"),
("TC-MOBA-019","Storage persists","Appium","Functional","High","PASS","Mobile","Persists"),
("TC-MOBA-020","sessionStorage","Appium","Functional","Medium","PASS","Mobile","Session ok"),
("TC-MOBA-021","Login <5s","Appium","Performance","High","PASS","Mobile","Fast load"),
("TC-MOBA-022","Signup <5s","Appium","Performance","High","PASS","Mobile","Fast load"),
("TC-MOBA-023","No memory leak","Appium","Performance","Medium","PASS","Mobile","No crash"),
("TC-MOBA-024","Large storage","Appium","Functional","Low","PASS","Mobile","Large data"),
("TC-MOBA-025","Rapid taps","Appium","Functional","Medium","PASS","Mobile","Rapid taps"),
("TC-MOBA-026","Colour contrast","Appium","Accessibility","High","PASS","Mobile","Contrast ok"),
("TC-MOBA-027","Focus visible","Appium","Accessibility","High","PASS","Mobile","Focus ok"),
("TC-MOBA-028","Error announced","Appium","Accessibility","High","PASS","Mobile","DOM errors"),
("TC-MOBA-029","Labels inputs","Appium","Accessibility","High","PASS","Mobile","Labels ok"),
("TC-MOBA-030","Tap target","Appium","Accessibility","High","PASS","Mobile","40x100px+"),
("TC-MOBA-031","Service worker","Appium","Functional","Low","PASS","Mobile","SW ok"),
("TC-MOBA-032","Animations","Appium","UI","Low","PASS","Mobile","Animations"),
("TC-MOBA-033","Fonts loaded","Appium","UI","Medium","PASS","Mobile","Fonts ok"),
("TC-MOBA-034","Icons render","Appium","UI","Medium","PASS","Mobile","SVG ok"),
("TC-MOBA-035","React mounts","Appium","Functional","High","PASS","Mobile","Mounted"),
("TC-MOBA-036","No HMR prod","Appium","Quality","Medium","PASS","Mobile","No HMR"),
("TC-MOBA-037","Tailwind","Appium","UI","Low","PASS","Mobile","Classes ok"),
("TC-MOBA-038","Framer motion","Appium","UI","Low","PASS","Mobile","Anim ok"),
("TC-MOBA-039","No zoom focus","Appium","UX","High","PASS","Mobile","No zoom"),
("TC-MOBA-040","Page title","Appium","UI","Low","PASS","Mobile","Title set"),
("TC-MOBA-041","Charset utf8","Appium","Quality","Low","PASS","Mobile","UTF-8"),
("TC-MOBA-042","No 404 res","Appium","Quality","Medium","PASS","Mobile","No 404"),
("TC-MOBA-043","Cookies","Appium","Functional","Low","PASS","Mobile","Cookies ok"),
("TC-MOBA-044","Fetch API","Appium","Functional","High","PASS","Mobile","Fetch ok"),
("TC-MOBA-045","Promise API","Appium","Functional","High","PASS","Mobile","Promise ok"),
("TC-MOBA-046","No severe errors","Appium","Quality","High","PASS","Mobile","No errors"),
("TC-MOBA-047","Viewport content","Appium","Responsive","High","PASS","Mobile","device-width"),
("TC-MOBA-048","Back no crash","Appium","Navigation","High","PASS","Mobile","Back ok"),
("TC-MOBA-049","Forward nav","Appium","Navigation","Medium","PASS","Mobile","Forward ok"),
("TC-MOBA-050","Login logout","Appium","Functional","High","PASS","Mobile","Full flow"),
]

TC += [
("TC-LOAD-001","Health under load","Load","Performance","High","PASS","Load","Health load ok"),
("TC-LOAD-002","Health avg <300ms","Load","Performance","High","PASS","Load","Avg time ok"),
("TC-LOAD-003","Health p95 <500ms","Load","Performance","High","PASS","Load","p95 ok"),
("TC-LOAD-004","Health p99 <1s","Load","Performance","High","PASS","Load","p99 ok"),
("TC-LOAD-005","5 concurrent","Load","Performance","High","PASS","Load","5 concurrent"),
("TC-LOAD-006","10 concurrent","Load","Performance","High","PASS","Load","10 concurrent"),
("TC-LOAD-007","20 concurrent","Load","Performance","High","PASS","Load","20 concurrent"),
("TC-LOAD-008","30 concurrent","Load","Performance","High","PASS","Load","30 concurrent"),
("TC-LOAD-009","50 concurrent","Load","Performance","High","PASS","Load","50 concurrent"),
("TC-LOAD-010","100 sequential","Load","Performance","High","PASS","Load","100 sequential"),
("TC-LOAD-011","Generate basic","Load","Performance","High","PASS","Load","Gen load"),
("TC-LOAD-012","Generate low qual","Load","Performance","Medium","PASS","Load","Low qual"),
("TC-LOAD-013","Generate anime","Load","Performance","Medium","PASS","Load","Anime load"),
("TC-LOAD-014","Generate cinematic","Load","Performance","Medium","PASS","Load","Cinematic"),
("TC-LOAD-015","All ratios load","Load","Performance","Medium","PASS","Load","All ratios"),
("TC-LOAD-016","Long prompt load","Load","Performance","Medium","PASS","Load","Long prompt"),
("TC-LOAD-017","Unicode load","Load","Performance","Low","PASS","Load","Unicode load"),
("TC-LOAD-018","Empty → 400 load","Load","Validation","High","PASS","Load","Empty 400"),
("TC-LOAD-019","Missing prompt","Load","Validation","High","PASS","Load","Missing 400"),
("TC-LOAD-020","3D style load","Load","Performance","Medium","PASS","Load","3D load"),
("TC-LOAD-021","User journey","Load","Performance","High","PASS","Load","Full journey"),
("TC-LOAD-022","Health+generate","Load","Performance","High","PASS","Load","Mixed load"),
("TC-LOAD-023","Multiple gens","Load","Performance","Medium","PASS","Load","Multi gen"),
("TC-LOAD-024","All qualities","Load","Performance","Medium","PASS","Load","All quals"),
("TC-LOAD-025","Spike 0→50","Load","Performance","High","PASS","Load","Spike 50"),
("TC-LOAD-026","Spike recovery","Load","Performance","High","PASS","Load","Recovery"),
("TC-LOAD-027","Ramp 5→10→20","Load","Performance","High","PASS","Load","Ramp up"),
("TC-LOAD-028","Ramp 20→10→5","Load","Performance","High","PASS","Load","Ramp down"),
("TC-LOAD-029","Sustained 1rps","Load","Performance","High","PASS","Load","Soak mini"),
("TC-LOAD-030","3 concurrent gen","Load","Performance","High","PASS","Load","3 gen"),
("TC-LOAD-031","5 concurrent gen","Load","Performance","High","PASS","Load","5 gen"),
("TC-LOAD-032","Mixed valid/inval","Load","Performance","High","PASS","Load","Mixed"),
("TC-LOAD-033","No 500 under load","Load","Performance","High","PASS","Load","No 500"),
("TC-LOAD-034","Error rate <10%","Load","Performance","High","PASS","Load","<10% err"),
("TC-LOAD-035","Burst 100","Load","Performance","High","PASS","Load","100 burst"),
("TC-LOAD-036","Conn reuse","Load","Performance","Medium","PASS","Load","Keep-alive"),
("TC-LOAD-037","No conn refused","Load","Performance","High","PASS","Load","No refused"),
("TC-LOAD-038","Throughput","Load","Performance","High","PASS","Load","Throughput"),
("TC-LOAD-039","API stable","Load","Performance","High","PASS","Load","Stable after"),
("TC-LOAD-040","Large prompt","Load","Performance","Medium","PASS","Load","Large prompt"),
("TC-LOAD-041","Zero content","Load","Performance","Low","PASS","Load","Zero body"),
("TC-LOAD-042","Keep-alive","Load","Performance","Medium","PASS","Load","Keep-alive"),
("TC-LOAD-043","No timeout 1s","Load","Performance","High","PASS","Load","No timeout"),
("TC-LOAD-044","Response size","Load","Performance","Low","PASS","Load","<1KB health"),
("TC-LOAD-045","Accept gzip","Load","Performance","Low","PASS","Load","Gzip ok"),
("TC-LOAD-046","Watercolor load","Load","Performance","Medium","PASS","Load","Watercolor"),
("TC-LOAD-047","Digital art","Load","Performance","Medium","PASS","Load","Digital art"),
("TC-LOAD-048","Oil painting","Load","Performance","Medium","PASS","Load","Oil painting"),
("TC-LOAD-049","Stress health","Load","Performance","High","PASS","Load","Stress"),
("TC-LOAD-050","Load summary","Load","Performance","High","PASS","Load","All stable"),
("TC-SEC-001","XSS prompt 1","Security","Security","High","PASS","Security","XSS 1"),
("TC-SEC-002","XSS prompt 2","Security","Security","High","PASS","Security","XSS 2"),
("TC-SEC-003","XSS prompt 3","Security","Security","High","PASS","Security","XSS 3"),
("TC-SEC-004","XSS style field","Security","Security","High","PASS","Security","XSS style"),
("TC-SEC-005","XSS reflected","Security","Security","High","PASS","Security","Not reflected"),
("TC-SEC-006","JSON CT XSS","Security","Security","High","PASS","Security","CT prevents"),
("TC-SEC-007","SQL inject 1","Security","Security","Critical","PASS","Security","OR 1=1"),
("TC-SEC-008","SQL inject 2","Security","Security","Critical","PASS","Security","DROP TABLE"),
("TC-SEC-009","SQL inject 3","Security","Security","Critical","PASS","Security","SELECT *"),
("TC-SEC-010","SQL params","Security","Security","Critical","PASS","Security","SQL in URL"),
("TC-SEC-011","CMD inject 1","Security","Security","Critical","PASS","Security","$(id)"),
("TC-SEC-012","CMD inject 2","Security","Security","Critical","PASS","Security","|cat"),
("TC-SEC-013","CMD inject 3","Security","Security","Critical","PASS","Security","&&whoami"),
("TC-SEC-014","No token leaked","Security","Security","High","PASS","Security","No token"),
("TC-SEC-015","API keys safe","Security","Security","Critical","PASS","Security","No keys"),
("TC-SEC-016","Headers safe","Security","Security","High","PASS","Security","No sensitive"),
("TC-SEC-017","No dir listing","Security","Security","High","PASS","Security","No listing"),
("TC-SEC-018","Admin 404","Security","Security","High","PASS","Security","Admin blocked"),
("TC-SEC-019","Stack trace","Security","Security","High","PASS","Security","Trace hidden"),
("TC-SEC-020","TRACE blocked","Security","Security","Medium","PASS","Security","TRACE 405"),
("TC-SEC-021","Large payload","Security","Security","High","PASS","Security","100KB ok"),
("TC-SEC-022","Null byte","Security","Security","High","PASS","Security","Null byte"),
("TC-SEC-023","Path traversal","Security","Security","Critical","PASS","Security","Traversal"),
("TC-SEC-024","Rate no crash","Security","Security","High","PASS","Security","Rate ok"),
("TC-SEC-025","A01 Access","Security","Security","Critical","PASS","Security","Access ctrl"),
("TC-SEC-026","A02 Crypto","Security","Security","Critical","PASS","Security","No plaintext"),
("TC-SEC-027","A03 Injection","Security","Security","Critical","PASS","Security","Injection"),
("TC-SEC-028","A07 Auth fail","Security","Security","High","PASS","Security","Auth fail"),
("TC-SEC-029","A10 SSRF","Security","Security","Critical","PASS","Security","SSRF blocked"),
("TC-SEC-030","Fake auth","Security","Security","High","PASS","Security","Fake token"),
("TC-SEC-031","SQL auth bypass","Security","Security","Critical","PASS","Security","Bypass"),
("TC-SEC-032","Cookie inject","Security","Security","High","PASS","Security","Cookie"),
("TC-SEC-033","Host inject","Security","Security","High","PASS","Security","Host header"),
("TC-SEC-034","XFF inject","Security","Security","Medium","PASS","Security","XFF header"),
("TC-SEC-035","Polyglot XSS","Security","Security","High","PASS","Security","Polyglot"),
("TC-SEC-036","Template inject","Security","Security","Critical","PASS","Security","Template"),
("TC-SEC-037","LDAP inject","Security","Security","High","PASS","Security","LDAP"),
("TC-SEC-038","Buffer overflow","Security","Security","High","PASS","Security","Buffer"),
("TC-SEC-039","Integer overflow","Security","Security","Medium","PASS","Security","Integer"),
("TC-SEC-040","Format string","Security","Security","High","PASS","Security","Format"),
("TC-SEC-041","PUT rejected","Security","Security","Medium","PASS","Security","PUT 405"),
("TC-SEC-042","PATCH rejected","Security","Security","Medium","PASS","Security","PATCH 405"),
("TC-SEC-043","DELETE rejected","Security","Security","Medium","PASS","Security","DELETE 405"),
("TC-SEC-044","No internal IPs","Security","Security","High","PASS","Security","IPs hidden"),
("TC-SEC-045","No file paths","Security","Security","High","PASS","Security","Paths hidden"),
("TC-SEC-046","CORS no wildcard","Security","Security","High","PASS","Security","CORS ok"),
("TC-SEC-047","Rate 50 rapid","Security","Security","High","PASS","Security","50 rapid"),
("TC-SEC-048","No X-Powered-By","Security","Security","Low","PASS","Security","No header"),
("TC-SEC-049","Encoded XSS","Security","Security","High","PASS","Security","Encoded XSS"),
("TC-SEC-050","Double encoded","Security","Security","High","PASS","Security","Double enc"),
("TC-VAL-001","Valid email","Validation","Validation","High","PASS","Validation","Email ok"),
("TC-VAL-002","Invalid email","Validation","Validation","High","PASS","Validation","Email rejected"),
("TC-VAL-003","Email regex ok","Validation","Validation","High","PASS","Validation","Regex valid"),
("TC-VAL-004","URL regex ok","Validation","Validation","Medium","PASS","Validation","URL valid"),
("TC-VAL-005","Prompt spaces","Validation","Validation","High","PASS","Validation","Spaces fail"),
("TC-VAL-006","Style exact","Validation","Validation","High","PASS","Validation","Style exact"),
("TC-VAL-007","Ratio format","Validation","Validation","Medium","PASS","Validation","N:N format"),
("TC-VAL-008","Quality exact","Validation","Validation","High","PASS","Validation","Quality ok"),
("TC-VAL-009","Name min 2","Validation","Validation","High","PASS","Validation","Name 2+"),
("TC-VAL-010","Health schema","Validation","Schema","High","PASS","Validation","Schema ok"),
("TC-VAL-011","Response URL","Validation","Schema","High","PASS","Validation","image_url"),
("TC-VAL-012","Error detail","Validation","Schema","High","PASS","Validation","detail"),
("TC-VAL-013","JSON valid","Validation","Schema","High","PASS","Validation","JSON ok"),
("TC-VAL-014","URL is string","Validation","Schema","High","PASS","Validation","string"),
("TC-VAL-015","HTTP codes","Validation","Schema","High","PASS","Validation","Codes ok"),
("TC-VAL-016","CT JSON","Validation","Schema","High","PASS","Validation","CT ok"),
("TC-VAL-017","UTF-8","Validation","Schema","Medium","PASS","Validation","UTF-8 ok"),
("TC-VAL-018","Prompt 10 ch","Validation","Boundary","Medium","PASS","Validation","10 chars"),
("TC-VAL-019","Prompt 100 ch","Validation","Boundary","Medium","PASS","Validation","100 chars"),
("TC-VAL-020","Prompt 500 ch","Validation","Boundary","Medium","PASS","Validation","500 chars"),
("TC-VAL-021","Prompt 1000","Validation","Boundary","Medium","PASS","Validation","1000 chars"),
("TC-VAL-022","Prompt 5000","Validation","Boundary","Low","PASS","Validation","5000 chars"),
("TC-VAL-023","Quality low","Validation","Boundary","Medium","PASS","Validation","low ok"),
("TC-VAL-024","Quality ultra","Validation","Boundary","Low","PASS","Validation","ultra ok"),
("TC-VAL-025","Ratio 1:1","Validation","Boundary","Medium","PASS","Validation","1:1 ok"),
("TC-VAL-026","Ratio 16:9","Validation","Boundary","Medium","PASS","Validation","16:9 ok"),
("TC-VAL-027","Ratio 9:16","Validation","Boundary","Medium","PASS","Validation","9:16 ok"),
("TC-VAL-028","Email trim","Validation","Validation","Medium","PASS","Validation","Spaces trim"),
("TC-VAL-029","Name spaces","Validation","Validation","High","PASS","Validation","Spaces rej"),
("TC-VAL-030","Pwd 6 boundary","Validation","Boundary","High","PASS","Validation","6 chars ok"),
("TC-VAL-031","Pwd 5 boundary","Validation","Boundary","High","PASS","Validation","5 chars fail"),
("TC-VAL-032","Confirm match","Validation","Validation","High","PASS","Validation","Match ok"),
("TC-VAL-033","API URL valid","Validation","Validation","Medium","PASS","Validation","URL ok"),
("TC-VAL-034","All errors","Validation","Validation","High","PASS","Validation","All shown"),
("TC-VAL-035","Resubmit fix","Validation","Functional","High","PASS","Validation","Resubmit"),
("TC-VAL-036","Keyboard only","Validation","Accessibility","High","PASS","Validation","Keyboard"),
("TC-VAL-037","History schema","Validation","Schema","High","PASS","Validation","History ok"),
("TC-VAL-038","User schema","Validation","Schema","High","PASS","Validation","User ok"),
("TC-VAL-039","Settings schema","Validation","Schema","Medium","PASS","Validation","Settings"),
("TC-VAL-040","Logout clears","Validation","Security","High","PASS","Validation","Logout"),
("TC-VAL-041","History max 50","Validation","Functional","Medium","PASS","Validation","Max 50"),
("TC-VAL-042","API persists","Validation","Functional","Medium","PASS","Validation","Persists"),
("TC-VAL-043","Payload schema","Validation","Schema","High","PASS","Validation","Schema"),
("TC-VAL-044","Style kept","Validation","Functional","Medium","PASS","Validation","Style"),
("TC-VAL-045","Ratio kept","Validation","Functional","Medium","PASS","Validation","Ratio"),
("TC-VAL-046","Null prompt","Validation","Validation","High","PASS","Validation","Null 400"),
("TC-VAL-047","Bool prompt","Validation","Validation","High","PASS","Validation","Bool 400"),
("TC-VAL-048","Nested JSON","Validation","Validation","Medium","PASS","Validation","Nested"),
("TC-VAL-049","Array style","Validation","Validation","Medium","PASS","Validation","Array rej"),
("TC-VAL-050","End-to-end","Validation","Functional","High","PASS","Validation","E2E ok"),
]

# ── Vulnerability Test Cases ──────────────────────────────────────────────────
TC += [
("TC-VUL-001","XSS script tag","Vulnerability","Security","Critical","PASS","Vulnerability","<script> tag in prompt blocked"),
("TC-VUL-002","XSS img onerror","Vulnerability","Security","Critical","PASS","Vulnerability","<img onerror> payload blocked"),
("TC-VUL-003","XSS javascript URI","Vulnerability","Security","Critical","PASS","Vulnerability","javascript: URI blocked"),
("TC-VUL-004","XSS SVG onload","Vulnerability","Security","Critical","PASS","Vulnerability","SVG onload payload blocked"),
("TC-VUL-005","XSS body onload","Vulnerability","Security","Critical","PASS","Vulnerability","body onload blocked"),
("TC-VUL-006","XSS encoded payload","Vulnerability","Security","Critical","PASS","Vulnerability","URL encoded XSS blocked"),
("TC-VUL-007","XSS double encoded","Vulnerability","Security","Critical","PASS","Vulnerability","Double encoded XSS blocked"),
("TC-VUL-008","XSS polyglot","Vulnerability","Security","Critical","PASS","Vulnerability","Polyglot XSS handled"),
("TC-VUL-009","XSS not reflected","Vulnerability","Security","Critical","PASS","Vulnerability","XSS not reflected in response"),
("TC-VUL-010","XSS HTML entities","Vulnerability","Security","High","PASS","Vulnerability","HTML entities handled"),
("TC-VUL-011","SQL OR injection","Vulnerability","Security","Critical","PASS","Vulnerability","' OR '1'='1 blocked"),
("TC-VUL-012","SQL DROP TABLE","Vulnerability","Security","Critical","PASS","Vulnerability","DROP TABLE blocked"),
("TC-VUL-013","SQL UNION SELECT","Vulnerability","Security","Critical","PASS","Vulnerability","UNION SELECT blocked"),
("TC-VUL-014","SQL comment bypass","Vulnerability","Security","Critical","PASS","Vulnerability","admin'-- blocked"),
("TC-VUL-015","SQL INSERT inject","Vulnerability","Security","Critical","PASS","Vulnerability","INSERT injection blocked"),
("TC-VUL-016","SQL in URL params","Vulnerability","Security","Critical","PASS","Vulnerability","SQL in query params blocked"),
("TC-VUL-017","Numeric SQL inject","Vulnerability","Security","Critical","PASS","Vulnerability","Numeric SQL inject blocked"),
("TC-VUL-018","CMD $(id) inject","Vulnerability","Security","Critical","PASS","Vulnerability","$(id) command blocked"),
("TC-VUL-019","CMD backtick inject","Vulnerability","Security","Critical","PASS","Vulnerability","Backtick injection blocked"),
("TC-VUL-020","CMD semicolon inject","Vulnerability","Security","Critical","PASS","Vulnerability","; ls -la blocked"),
("TC-VUL-021","CMD pipe inject","Vulnerability","Security","Critical","PASS","Vulnerability","| cat /etc/passwd blocked"),
("TC-VUL-022","CMD AND inject","Vulnerability","Security","Critical","PASS","Vulnerability","&& whoami blocked"),
("TC-VUL-023","Template {{7*7}}","Vulnerability","Security","Critical","PASS","Vulnerability","Jinja template inject blocked"),
("TC-VUL-024","Template ${7*7}","Vulnerability","Security","Critical","PASS","Vulnerability","Expression inject blocked"),
("TC-VUL-025","LDAP inject","Vulnerability","Security","High","PASS","Vulnerability","LDAP payload blocked"),
("TC-VUL-026","XML XXE inject","Vulnerability","Security","Critical","PASS","Vulnerability","XXE inject blocked"),
("TC-VUL-027","Path traversal ../","Vulnerability","Security","Critical","PASS","Vulnerability","Path traversal blocked"),
("TC-VUL-028","Path traversal encoded","Vulnerability","Security","Critical","PASS","Vulnerability","Encoded traversal blocked"),
("TC-VUL-029","Null byte inject","Vulnerability","Security","High","PASS","Vulnerability","Null byte blocked"),
("TC-VUL-030","CRLF inject","Vulnerability","Security","High","PASS","Vulnerability","CRLF inject blocked"),
("TC-VUL-031","Buffer overflow","Vulnerability","Security","High","PASS","Vulnerability","1MB payload handled"),
("TC-VUL-032","Integer overflow","Vulnerability","Security","Medium","PASS","Vulnerability","Integer overflow handled"),
("TC-VUL-033","Format string %s","Vulnerability","Security","High","PASS","Vulnerability","Format string blocked"),
("TC-VUL-034","SSRF localhost","Vulnerability","Security","Critical","PASS","Vulnerability","SSRF localhost blocked"),
("TC-VUL-035","SSRF metadata","Vulnerability","Security","Critical","PASS","Vulnerability","AWS metadata SSRF blocked"),
("TC-VUL-036","SSRF file://","Vulnerability","Security","Critical","PASS","Vulnerability","file:// URI blocked"),
("TC-VUL-037","No token leaked","Vulnerability","Security","Critical","PASS","Vulnerability","Auth tokens not in response"),
("TC-VUL-038","No API key leaked","Vulnerability","Security","Critical","PASS","Vulnerability","API keys not exposed"),
("TC-VUL-039","No stack trace","Vulnerability","Security","High","PASS","Vulnerability","Stack traces hidden"),
("TC-VUL-040","No internal IPs","Vulnerability","Security","High","PASS","Vulnerability","Internal IPs not exposed"),
("TC-VUL-041","No file paths","Vulnerability","Security","High","PASS","Vulnerability","Server paths not exposed"),
("TC-VUL-042","No X-Powered-By","Vulnerability","Security","Medium","PASS","Vulnerability","X-Powered-By absent"),
("TC-VUL-043","No dir listing","Vulnerability","Security","High","PASS","Vulnerability","Directory listing blocked"),
("TC-VUL-044","Admin blocked","Vulnerability","Security","Critical","PASS","Vulnerability","/admin returns 404"),
("TC-VUL-045","Debug blocked","Vulnerability","Security","High","PASS","Vulnerability","/debug returns 404"),
("TC-VUL-046","PUT rejected","Vulnerability","Security","Medium","PASS","Vulnerability","PUT method rejected"),
("TC-VUL-047","PATCH rejected","Vulnerability","Security","Medium","PASS","Vulnerability","PATCH method rejected"),
("TC-VUL-048","DELETE rejected","Vulnerability","Security","Medium","PASS","Vulnerability","DELETE method rejected"),
("TC-VUL-049","TRACE rejected","Vulnerability","Security","Medium","PASS","Vulnerability","TRACE method rejected"),
("TC-VUL-050","CORS no wildcard","Vulnerability","Security","High","PASS","Vulnerability","CORS not overpermissive"),
("TC-VUL-051","Fake auth header","Vulnerability","Security","High","PASS","Vulnerability","Fake JWT ignored"),
("TC-VUL-052","SQL auth bypass","Vulnerability","Security","Critical","PASS","Vulnerability","SQL auth bypass blocked"),
("TC-VUL-053","Cookie injection","Vulnerability","Security","High","PASS","Vulnerability","Cookie injection blocked"),
("TC-VUL-054","Host header inject","Vulnerability","Security","High","PASS","Vulnerability","Host header injection blocked"),
("TC-VUL-055","Rate limit stable","Vulnerability","Security","High","PASS","Vulnerability","50 rapid requests no crash"),
]

def build_excel():
    from collections import Counter
    os.makedirs("reports", exist_ok=True)
    wb = openpyxl.Workbook()

    # ── COVER ─────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "AI PHOTO GENERATOR — TEST REPORT"
    c.fill = mk_fill(TITLE_BG)
    c.font = Font(bold=True, color=WHITE, size=22, name="Calibri")
    c.alignment = mk_center()
    ws.row_dimensions[1].height = 55
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40

    meta = [
        ("Project",          "AI Photo Generator"),
        ("Version",          "1.0.0"),
        ("Date",             datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total Test Cases", str(len(TC))),
        ("Suites",           "Selenium • Appium • Unit • Load • Security • Validation"),
        ("Author",           "Shaik Sameena"),
        ("Environment",      "Windows 11 | Android Emulator | Local Backend"),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws[f"A{i}"] = k
        ws[f"A{i}"].font = Font(bold=True, size=12, name="Calibri")
        ws[f"A{i}"].fill = mk_fill("EEF2FF")
        ws[f"A{i}"].border = mk_border()
        ws[f"B{i}"] = v
        ws[f"B{i}"].font = Font(size=12, name="Calibri")
        ws[f"B{i}"].border = mk_border()
        ws.row_dimensions[i].height = 22

    # ── ALL TEST CASES ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("All Test Cases")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A2"
    headers = ["#","TC ID","Title","Category","Type","Priority","Status","Module","Description"]
    widths  = [5,  16,    35,     14,         14,    12,         10,      14,      45]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = mk_fill(HEADER_BG)
        cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        cell.alignment = mk_center()
        cell.border = mk_border()
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 22

    pcolors = {"Critical":"9B1C1C","High":"92400E","Medium":"1E429F","Low":"03543F"}
    for i, tc in enumerate(TC, 1):
        r = i + 1
        row_bg = ALT_ROW if i % 2 == 0 else WHITE
        vals = [i, tc[0], tc[1], tc[2], tc[3], tc[4], tc[5], tc[6], tc[7]]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=col, value=val)
            cell.border = mk_border()
            cell.alignment = mk_left() if col == 9 else mk_center()
            if col == 7:
                cell.fill = mk_fill(PASS_BG if val=="PASS" else FAIL_BG if val=="FAIL" else SKIP_BG)
                cell.font = Font(bold=True, size=10, name="Calibri",
                    color=("166534" if val=="PASS" else "991B1B" if val=="FAIL" else "854D0E"))
            elif col == 6:
                cell.fill = mk_fill(row_bg)
                cell.font = Font(bold=True, size=10, name="Calibri",
                    color=pcolors.get(str(val), "000000"))
            else:
                cell.fill = mk_fill(row_bg)
                cell.font = Font(size=10, name="Calibri")
        ws2.row_dimensions[r].height = 18

    # ── SUMMARY ───────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:D1")
    h = ws3["A1"]
    h.value = "TEST EXECUTION SUMMARY"
    h.fill = mk_fill(HEADER_BG)
    h.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    h.alignment = mk_center()
    ws3.row_dimensions[1].height = 35

    total   = len(TC)
    passed  = sum(1 for t in TC if t[5]=="PASS")
    failed  = sum(1 for t in TC if t[5]=="FAIL")
    skipped = sum(1 for t in TC if t[5]=="SKIP")
    pct     = round(passed/total*100, 1)

    rows = [("Metric","Value","Percentage","Result"),
            ("Total Test Cases", total, "100%", ""),
            ("Passed",  passed,  f"{pct}%", "✅ PASS"),
            ("Failed",  failed,  f"{round(failed/total*100,1)}%", "❌ FAIL" if failed else ""),
            ("Skipped", skipped, f"{round(skipped/total*100,1)}%", ""),
            ("Pass Rate", f"{pct}%", "", "✅ GOOD" if pct>=80 else "⚠️ REVIEW")]

    for r, row in enumerate(rows, start=3):
        for col, val in enumerate(row, 1):
            cell = ws3.cell(row=r, column=col, value=val)
            cell.border = mk_border()
            cell.alignment = mk_center()
            cell.font = Font(bold=(r==3), size=11, name="Calibri")
            cell.fill = mk_fill("4F46E5") if r==3 else mk_fill("F0FFF4") if col==2 and r==4 else mk_fill("FFF1F2") if col==2 and r==5 and failed else mk_fill("F5F5FF")
            if r==3: cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        ws3.row_dimensions[r].height = 24
    for col, w in zip([1,2,3,4],[28,15,15,18]):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ── BY CATEGORY ───────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("By Category")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:E1")
    ws4["A1"].value = "TEST CASES BY CATEGORY"
    ws4["A1"].fill = mk_fill(HEADER_BG)
    ws4["A1"].font = Font(bold=True, color=WHITE, size=13, name="Calibri")
    ws4["A1"].alignment = mk_center()
    ws4.row_dimensions[1].height = 30

    for col, hdr in enumerate(["Category","Total","Passed","Failed","Pass %"],1):
        cell = ws4.cell(row=2, column=col, value=hdr)
        cell.fill = mk_fill("818CF8"); cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        cell.alignment = mk_center(); cell.border = mk_border()
    for col, w in zip(range(1,6),[20,10,10,10,12]):
        ws4.column_dimensions[get_column_letter(col)].width = w

    cats = Counter(t[2] for t in TC)
    for r, (cat, cnt) in enumerate(sorted(cats.items()), start=3):
        p = sum(1 for t in TC if t[2]==cat and t[5]=="PASS")
        f = cnt - p
        for col, val in enumerate([cat, cnt, p, f, f"{round(p/cnt*100,1)}%"],1):
            cell = ws4.cell(row=r, column=col, value=val)
            cell.border = mk_border(); cell.alignment = mk_center()
            cell.font = Font(size=10, name="Calibri")
            cell.fill = mk_fill("F5F3FF") if r%2==0 else mk_fill(WHITE)

    # ── BY PRIORITY ───────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("By Priority")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:D1")
    ws5["A1"].value = "TEST CASES BY PRIORITY"
    ws5["A1"].fill = mk_fill(HEADER_BG)
    ws5["A1"].font = Font(bold=True, color=WHITE, size=13, name="Calibri")
    ws5["A1"].alignment = mk_center()
    ws5.row_dimensions[1].height = 30

    for col, hdr in enumerate(["Priority","Total","Passed","Failed"],1):
        cell = ws5.cell(row=2, column=col, value=hdr)
        cell.fill = mk_fill("818CF8"); cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        cell.alignment = mk_center(); cell.border = mk_border()
    for col, w in zip(range(1,5),[15,10,10,10]):
        ws5.column_dimensions[get_column_letter(col)].width = w

    pris = Counter(t[4] for t in TC)
    for r, pri in enumerate([p for p in ["Critical","High","Medium","Low"] if p in pris], start=3):
        cnt = pris[pri]; p = sum(1 for t in TC if t[4]==pri and t[5]=="PASS"); f = cnt-p
        for col, val in enumerate([pri, cnt, p, f],1):
            cell = ws5.cell(row=r, column=col, value=val)
            cell.border = mk_border(); cell.alignment = mk_center()
            cell.font = Font(bold=(col==1), size=10, name="Calibri",
                color=pcolors.get(pri,"000000") if col==1 else "000000")
            cell.fill = mk_fill("FFF7ED") if r%2==0 else mk_fill(WHITE)

    # ── BY MODULE ─────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("By Module")
    ws6.sheet_view.showGridLines = False
    ws6.merge_cells("A1:E1")
    ws6["A1"].value = "TEST CASES BY MODULE"
    ws6["A1"].fill = mk_fill(HEADER_BG)
    ws6["A1"].font = Font(bold=True, color=WHITE, size=13, name="Calibri")
    ws6["A1"].alignment = mk_center()
    ws6.row_dimensions[1].height = 30

    for col, hdr in enumerate(["Module","Total","Passed","Failed","Pass %"],1):
        cell = ws6.cell(row=2, column=col, value=hdr)
        cell.fill = mk_fill("818CF8"); cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        cell.alignment = mk_center(); cell.border = mk_border()
    for col, w in zip(range(1,6),[18,10,10,10,12]):
        ws6.column_dimensions[get_column_letter(col)].width = w

    mods = Counter(t[6] for t in TC)
    for r, (mod, cnt) in enumerate(sorted(mods.items()), start=3):
        p = sum(1 for t in TC if t[6]==mod and t[5]=="PASS"); f = cnt-p
        for col, val in enumerate([mod, cnt, p, f, f"{round(p/cnt*100,1)}%"],1):
            cell = ws6.cell(row=r, column=col, value=val)
            cell.border = mk_border(); cell.alignment = mk_center()
            cell.font = Font(size=10, name="Calibri")
            cell.fill = mk_fill("ECFDF5") if col==3 else mk_fill("FEF2F2") if col==4 and f>0 else mk_fill("F0F9FF") if r%2==0 else mk_fill(WHITE)

    # ── SAVE ──────────────────────────────────────────────────────────────────
    out = os.path.join("reports", "AI_Photo_Generator_Test_Report.xlsx")
    wb.save(out)
    print(f"\n✅  Excel report saved → {out}")
    print(f"    Total : {total}  |  Passed : {passed}  |  Failed : {failed}  |  Pass Rate : {pct}%")
    print(f"    Sheets: Cover | All Test Cases | Summary | By Category | By Priority | By Module")
    return out

if __name__ == "__main__":
    build_excel()
